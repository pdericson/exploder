#!/usr/bin/env python3

"""Exploder."""

import argparse
import copy
import itertools
import os
import sys

try:
    import openpyxl
except ModuleNotFoundError:
    sys.path.append(os.path.join(os.path.dirname(__file__), 'vendor'))
    import openpyxl


def explode(wb, ws1, ws2, cols):
    """Explode worksheet 1."""
    unique = {}
    style = {}
    style_row = 1
    row_dimension = None
    row_dimension_row = 1

    data = []
    for row in ws1.iter_rows():
        data.append([])
        for cell in row:
            if cell.column in cols and cell.value is not None:
                value = cell.value
                if isinstance(value, int):
                    value = str(value)
                data[-1].append([item.strip() for item in value.split(',')])
                unique.setdefault(cell.column, set()).update(data[-1][-1])
            else:
                data[-1].append([cell.value])
            style[cell.column] = copy.copy(cell._style)
            style_row = cell.row
            row_dimension = copy.copy(ws1.row_dimensions[cell.row])
            row_dimension_row = cell.row

    print(f"Worksheet {ws1.title} has {len(data)} rows.")
    for key in sorted(unique):
        print(f"Column {key} has the following unique items:")
        for val in sorted(unique[key]):
            print(f"- {val}")

    row = 1
    for item1 in data:
        for item2 in itertools.product(*item1):
            for column, value in enumerate(item2, 1):
                cell = ws2.cell(row=row, column=column, value=value)
                if row > style_row:
                    cell._style = copy.copy(style[column])
                if row > row_dimension_row:
                    ws2.row_dimensions[cell.row] = copy.copy(row_dimension)
            row += 1

    print(f"Worksheet {ws2.title} has {row - 1} rows.")


def main():
    parser = argparse.ArgumentParser(description="Exploder.")
    parser.add_argument('--worksheet1', metavar='WORKSHEET', required=True, help='worksheet 1')
    parser.add_argument('--worksheet2', metavar='WORKSHEET', required=True, help='worksheet 2')
    parser.add_argument('--columns', required=True, help='the columns to explode')
    parser.add_argument('path', help='the workbook path')
    args = parser.parse_args()

    args.columns = [int(column) for column in args.columns.split(',')]

    wb = openpyxl.load_workbook(filename=args.path)

    ws1 = wb[args.worksheet1]

    try:
        ws2 = wb[args.worksheet2]
    except KeyError:
        ws2 = wb.copy_worksheet(ws1)
        ws2.title = args.worksheet2

    explode(wb, ws1, ws2, args.columns)

    wb.save(args.path)


if __name__ == '__main__':
    main()
